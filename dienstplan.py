#!/usr/bin/env python3
"""
Dienstplan-Kalender
Extrahiert Dienste aus SP Expert PDF oder Excel-Vordienstplan und erstellt .ics Dateien.

Verwendung:
    python dienstplan.py dienstplan.pdf
    python dienstplan.py vordienstplan.xlsx
    python dienstplan.py dienstplan.pdf "Nachname, Vorname" --alle
    python dienstplan.py dienstplan.pdf --debug
"""

import sys
import re
import uuid
import json
import yaml
import pdfplumber
import openpyxl
import datetime as dt
import subprocess
import socket
import threading
import http.server
from datetime import datetime, timedelta, date
import pytz

BERLIN = pytz.timezone('Europe/Berlin')
from icalendar import Calendar, Event
from pathlib import Path


MONTH_MAP = {
    'Januar': 1, 'Februar': 2, 'März': 3, 'April': 4,
    'Mai': 5, 'Juni': 6, 'Juli': 7, 'August': 8,
    'September': 9, 'Oktober': 10, 'November': 11, 'Dezember': 12
}

# Kurzformen wie in der Excel-Datei: "Mrz 26"
MONTH_ABBR_MAP = {
    'Jan': 1, 'Feb': 2, 'Mrz': 3, 'Mär': 3, 'Mar': 3,
    'Apr': 4, 'Mai': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8,
    'Sep': 9, 'Okt': 10, 'Nov': 11, 'Dez': 12
}

EXCEL_ERRORS = {'#NV', '#N/A', '#WERT!', '#VALUE!', '#REF!', '#BEZUG!', '#DIV/0!', '#NAME?'}

WEEKEND = {'Sam', 'Son'}
WEEKDAY_NAMES = ['Mon', 'Die', 'Mit', 'Don', 'Fre', 'Sam', 'Son']


def dedup(s):
    """Entfernt aufeinanderfolgende doppelte Zeichen: 'JJuullii' -> 'Juli', 'MMiitt' -> 'Mit'"""
    if not s:
        return s
    result = [s[0]]
    for c in s[1:]:
        if c != result[-1]:
            result.append(c)
    return ''.join(result)


def load_config(config_path: Path) -> dict:
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def parse_time(time_str: str) -> tuple:
    h, m = time_str.split(':')
    return int(h), int(m)


def make_datetime(d: date, time_str: str) -> datetime:
    h, m = parse_time(time_str)
    return BERLIN.localize(datetime(d.year, d.month, d.day, h, m))


def extract_schedule(pdf_path: Path, name: str, debug: bool = False) -> tuple:
    """
    Gibt (schedule, month, year) zurück.
    schedule = {day_number: {'shift': 'KRD', 'weekday': 'Mon'}}

    SP Expert PDFs haben zwei überlagerte Textebenen (Original + Update), dadurch:
    - Header-Zellen haben doppelte Zeichen: 'JJuullii 220022', '11\nMMiitt'
    - Datenzellen können Original\nUpdate enthalten (letzter Wert = aktuell)
    """
    schedule = {}
    found_month = None
    found_year = None

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()

            if debug:
                print(f"\n--- Seite {page_num + 1}: {len(tables)} Tabelle(n) ---")

            for table in tables:
                if not table or len(table) < 2:
                    continue

                # Zeile 0 ist immer der Header
                row0 = table[0] or []
                if not row0:
                    continue

                # Doppelte Zeichen im Header-Monat entfernen: 'JJuullii 220022' -> 'Juli 2026'
                first_dedup = dedup(str(row0[0] or '').strip())
                m = re.search(r'(\w+)\s+(\d{4})', first_dedup)
                if not m or m.group(1) not in MONTH_MAP:
                    if debug:
                        print(f"  Kein Monatsheader erkannt: '{first_dedup[:30]}'")
                    continue

                found_month = MONTH_MAP[m.group(1)]
                found_year = int(m.group(2))

                # Wochentage aus Header-Zellen lesen: '11\nMMiitt' -> Tag 1, Wochentag 'Mit'
                # Spalte j entspricht Tag j (Spalte 1 = Tag 1, Spalte 2 = Tag 2, ...)
                col_to_day = {}
                col_to_weekday = {}

                for j, cell in enumerate(row0[1:], 1):
                    cell_str = str(cell or '').strip()
                    if not cell_str:
                        continue
                    parts = cell_str.split('\n')
                    # Erste Zeile enthält die (doppelte) Tageszahl - Spaltenindex = Tagesnummer
                    day_part = dedup(parts[0].strip())
                    if re.search(r'\d', day_part):
                        col_to_day[j] = j  # Spalte 1 = Tag 1, Spalte 2 = Tag 2 usw.
                    # Zweite Zeile enthält den (doppelten) Wochentag
                    if len(parts) > 1:
                        wd = dedup(parts[1].strip())
                        if wd in {'Mon', 'Die', 'Mit', 'Don', 'Fre', 'Sam', 'Son'}:
                            col_to_weekday[j] = wd

                if not col_to_day:
                    if debug:
                        print("  Keine Tagesspalten gefunden")
                    continue

                if debug:
                    print(f"  Header: {m.group(1)} {found_year}, {len(col_to_day)} Tage, {len(col_to_weekday)} Wochentage")
                    print(f"  Erste 6 Zeilen:")
                    for i, row in enumerate(table[:6]):
                        cells = [str(c or '').replace('\n', '|')[:18] for c in (row or [])[:9]]
                        print(f"    Zeile {i}: {cells}")

                # Name in Datenzeilen suchen
                for i in range(1, len(table)):
                    row = table[i] or []
                    first = str(row[0] if row else '').strip()

                    if name.lower() not in first.lower():
                        continue

                    if debug:
                        cells = [str(c or '').replace('\n', '|')[:18] for c in row[:9]]
                        print(f"  NAME gefunden Zeile {i}: {cells}")

                    # Nächste Zeile prüfen: leere erste Spalte = Aktualisierungszeile
                    active_row = row
                    if i + 1 < len(table):
                        next_row = table[i + 1] or []
                        next_first = str(next_row[0] if next_row else '').strip()
                        if not next_first:
                            active_row = next_row
                            if debug:
                                cells2 = [str(c or '').replace('\n', '|')[:18] for c in next_row[:9]]
                                print(f"  UPDATE Zeile {i+1}: {cells2}")

                    for col, day in col_to_day.items():
                        if col >= len(active_row):
                            continue
                        cell_val = str(active_row[col] or '').strip()
                        # Bei mehreren Werten (Original\nUpdate) den letzten nehmen
                        if '\n' in cell_val:
                            parts = [p.strip() for p in cell_val.split('\n') if p.strip()]
                            cell_val = parts[-1] if parts else ''
                        if cell_val:
                            schedule[day] = {
                                'shift': cell_val,
                                'weekday': col_to_weekday.get(col, '')
                            }
                    break  # Person auf dieser Seite gefunden

    return schedule, found_month, found_year


def build_event(d: date, name: str, start_str: str, end_str: str, next_day: bool) -> Event:
    event = Event()
    event.add('summary', name)
    event.add('dtstart', make_datetime(d, start_str))
    end_date = d + timedelta(days=1) if next_day else d
    event.add('dtend', make_datetime(end_date, end_str))
    event.add('dtstamp', datetime.utcnow())
    event.add('uid', str(uuid.uuid4()))
    return event


def create_events(day: int, month: int, year: int, shift_info: dict, config: dict) -> list:
    shift_key = shift_info['shift']
    is_weekend = shift_info.get('weekday', '') in WEEKEND
    d = date(year, month, day)
    shifts = config.get('shifts', {})

    if shift_key not in shifts:
        # Unbekanntes Kürzel → ganztägiger Eintrag
        event = Event()
        event.add('summary', f'Dienst: {shift_key}')
        event.add('dtstart', d)
        event.add('dtend', d + timedelta(days=1))
        event.add('dtstamp', datetime.utcnow())
        event.add('uid', str(uuid.uuid4()))
        return [event]

    defn = shifts[shift_key]

    if defn.get('skip'):
        return []

    # k+K und ähnliche: mehrere Einträge
    if 'entries' in defn:
        return [
            build_event(d, e['name'], e['start'], e['end'], e.get('next_day', False))
            for e in defn['entries']
        ]

    # KRD und ähnliche: Wochentag vs. Wochenende
    if 'weekday' in defn and 'weekend' in defn:
        v = defn['weekend'] if is_weekend else defn['weekday']
        return [build_event(d, defn['name'], v['start'], v['end'], v.get('next_day', False))]

    # Normaler Dienst
    return [build_event(d, defn['name'], defn['start'], defn['end'], defn.get('next_day', False))]


def scan_excel(excel_path: Path):
    """Zeigt rohe Zellwerte der ersten 12 Zeilen und Spalten A-P zur Diagnose."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    print(f"\nAktives Tabellenblatt: '{ws.title}'")
    print(f"Blätter: {wb.sheetnames}\n")
    print("Erste 12 Zeilen (Wert | Typ):\n")
    for r in range(1, min(13, ws.max_row + 1)):
        row_info = []
        for c in range(1, min(18, ws.max_column + 1)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_info.append(f"[{openpyxl.utils.get_column_letter(c)}{r}] {repr(val)[:25]}")
        if row_info:
            print(f"  Zeile {r:2d}: " + "  |  ".join(row_info[:6]))
        else:
            print(f"  Zeile {r:2d}: (leer)")
    print()


def extract_schedule_excel(excel_path: Path, name: str, debug: bool = False) -> tuple:
    """
    Extrahiert Dienste aus der Excel-Vordienstplan-Datei.

    Besonderheiten dieses Formats:
    - Tageszahlen sind datetime-Objekte (datetime(2026, 3, 1) etc.), keine Zahlen
    - Namen stehen nur als Nachname in Spalte A (kein Komma, kein Vorname)
    - Monat/Jahr steht als datetime in Zeile 2
    - Kontrollspalten/-zeilen am Rand werden per Lückenerkennung ignoriert
    """
    import warnings
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    found_month = None
    found_year = None

    def cell_as_day(val):
        """Gibt Tageszahl zurück wenn val ein Datum oder eine Zahl 1-31 ist, sonst None."""
        if isinstance(val, (dt.datetime, dt.date)):
            return val.day if 1 <= val.day <= 31 else None
        try:
            n = int(str(val).strip())
            return n if 1 <= n <= 31 else None
        except (ValueError, TypeError):
            return None

    # 1. Monat/Jahr: datetime-Objekte in den ersten Zeilen auswerten
    for r in range(1, min(8, max_row + 1)):
        for c in range(1, min(max_col + 1, 30)):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, (dt.datetime, dt.date)):
                found_month = val.month
                found_year  = val.year
                break
            if isinstance(val, str):
                m = re.match(r'^(\w{3})\s+(\d{2,4})$', val.strip())
                if m and m.group(1) in MONTH_ABBR_MAP:
                    found_month = MONTH_ABBR_MAP[m.group(1)]
                    y = int(m.group(2))
                    found_year  = y + 2000 if y < 100 else y
                    break
                m2 = re.search(r'(\w+)\s+(\d{4})', val.strip())
                if m2 and m2.group(1) in MONTH_MAP:
                    found_month = MONTH_MAP[m2.group(1)]
                    found_year  = int(m2.group(2))
                    break
        if found_month:
            break

    # 2. Kopfzeile suchen: Zeile mit mind. 20 Datumswerten (Tag 1–31)
    header_row = None
    col_to_day  = {}

    for r in range(1, min(15, max_row + 1)):
        temp = {}
        for c in range(1, max_col + 1):
            d_num = cell_as_day(ws.cell(row=r, column=c).value)
            if d_num is not None:
                temp[c] = d_num
                # Monat/Jahr aus Datumszelle holen falls noch nicht bekannt
                if not found_month:
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, (dt.datetime, dt.date)):
                        found_month = val.month
                        found_year  = val.year

        if len(temp) >= 20:
            header_row = r
            # Zusammenhängende Tagesspalten (Lücke > 3 = Kontrollspalten)
            for i, c in enumerate(sorted(temp.keys())):
                prev = sorted(temp.keys())[i - 1] if i > 0 else c
                if i > 0 and c - prev > 3:
                    break
                col_to_day[c] = temp[c]
            if debug:
                print(f"  Kopfzeile: Zeile {header_row}, {len(col_to_day)} Tage "
                      f"(Spalten {openpyxl.utils.get_column_letter(min(col_to_day))}"
                      f"–{openpyxl.utils.get_column_letter(max(col_to_day))})")
            break

    if not header_row or not col_to_day:
        if debug:
            print("  Keine Tages-Kopfzeile gefunden.")
        return {}, found_month, found_year

    # 3. Namensspalte: linkeste Spalte mit Textwerten (Namen ohne Komma möglich)
    first_day_col = min(col_to_day.keys())
    name_col = None

    for c in range(1, first_day_col):
        hits = sum(
            1 for r in range(header_row + 1, min(header_row + 25, max_row + 1))
            if isinstance(ws.cell(row=r, column=c).value, str)
            and len(ws.cell(row=r, column=c).value.strip()) >= 2
        )
        if hits >= 5:
            name_col = c
            break

    if name_col is None:
        if debug:
            print("  Namensspalte nicht gefunden.")
        return {}, found_month, found_year

    if debug:
        print(f"  Namensspalte: {openpyxl.utils.get_column_letter(name_col)}")

    # 4. Person suchen und Dienste auslesen
    schedule   = {}
    row_limit  = min(max_row, header_row + 100)

    for r in range(header_row + 1, row_limit + 1):
        cell_val = ws.cell(row=r, column=name_col).value
        if not isinstance(cell_val, str):
            continue
        if name.lower() not in cell_val.strip().lower():
            continue

        if debug:
            print(f"  Name gefunden: Zeile {r} → '{cell_val.strip()}'")

        for col, day in col_to_day.items():
            shift_val = ws.cell(row=r, column=col).value
            if shift_val is None:
                continue
            shift_str = str(shift_val).strip()
            if not shift_str or shift_str in EXCEL_ERRORS:
                continue
            wd = ''
            if found_month and found_year:
                try:
                    wd = WEEKDAY_NAMES[date(found_year, found_month, day).weekday()]
                except ValueError:
                    pass
            schedule[day] = {'shift': shift_str, 'weekday': wd}
        break

    if debug and not schedule:
        print(f"  '{name}' nicht gefunden. Vorhandene Namen (erste 10):")
        shown = 0
        for r in range(header_row + 1, row_limit + 1):
            val = ws.cell(row=r, column=name_col).value
            if isinstance(val, str) and len(val.strip()) >= 2:
                print(f"    Zeile {r}: '{val.strip()}'")
                shown += 1
                if shown >= 10:
                    break

    return schedule, found_month, found_year


def find_all_names_excel(excel_path: Path) -> list:
    """Gibt alle Namen aus der Excel-Datei zurück."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws  = wb.active

    def cell_as_day(val):
        if isinstance(val, (dt.datetime, dt.date)):
            return val.day if 1 <= val.day <= 31 else None
        try:
            n = int(str(val).strip())
            return n if 1 <= n <= 31 else None
        except (ValueError, TypeError):
            return None

    # Kopfzeile finden
    header_row    = None
    first_day_col = None
    for r in range(1, min(15, ws.max_row + 1)):
        count = sum(1 for c in range(1, ws.max_column + 1)
                    if cell_as_day(ws.cell(row=r, column=c).value) is not None)
        if count >= 20:
            header_row    = r
            first_day_col = next(c for c in range(1, ws.max_column + 1)
                                 if cell_as_day(ws.cell(row=r, column=c).value) is not None)
            break

    if not header_row:
        return []

    # Namensspalte finden
    name_col = None
    for c in range(1, first_day_col):
        hits = sum(1 for r in range(header_row + 1, min(header_row + 25, ws.max_row + 1))
                   if isinstance(ws.cell(row=r, column=c).value, str)
                   and len(ws.cell(row=r, column=c).value.strip()) >= 2)
        if hits >= 5:
            name_col = c
            break

    if not name_col:
        return []

    names = []
    seen  = set()
    for r in range(header_row + 1, min(ws.max_row + 1, header_row + 100)):
        val = ws.cell(row=r, column=name_col).value
        if isinstance(val, str) and len(val.strip()) >= 2:
            n = val.strip()
            if n not in seen:
                seen.add(n)
                names.append(n)

    return sorted(names)


def _is_day_num(val) -> bool:
    try:
        n = int(str(val).strip())
        return 1 <= n <= 31
    except (ValueError, TypeError):
        return False


def find_all_names(pdf_path: Path) -> list:
    """Gibt alle Namen zurück die im PDF als Personen-Zeilen vorkommen."""
    names = []
    seen = set()

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                if not table or len(table) < 2:
                    continue
                # Header-Erkennung: erste Zelle muss Monat+Jahr enthalten
                first_dedup = dedup(str((table[0] or [''])[0] or '').strip())
                if not re.search(r'\w+\s+\d{4}', first_dedup):
                    continue
                # Alle Zeilen ab Index 1 mit nicht-leerem Namensfeld
                for row in table[1:]:
                    if not row:
                        continue
                    name = str(row[0] or '').strip()
                    # Namen haben typischerweise ein Komma (Nachname, Vorname)
                    if name and ',' in name and name not in seen:
                        seen.add(name)
                        names.append(name)
    return sorted(names)


def find_unknown_shifts(schedule: dict, config: dict) -> list:
    known = set(config.get('shifts', {}).keys())
    seen, unknown = set(), []
    for info in schedule.values():
        key = info['shift']
        if key not in known and key not in seen:
            unknown.append(key)
            seen.add(key)
    return unknown


def ask_and_save_shift(shift_key: str, config_path: Path, config: dict):
    """Fragt interaktiv nach Zeiten für ein unbekanntes Kürzel und speichert es in config.yaml."""
    print(f"\n  Unbekanntes Kuerzel gefunden: '{shift_key}'")
    print("  ----------------------------------------")

    create = input("  Kalendereintrag erstellen? (j/n) [j]: ").strip().lower() or 'j'

    if create != 'j':
        new_def = {'skip': True}
        yaml_block = f'\n  "{shift_key}":\n    skip: true\n'
        label = "(kein Eintrag)"
    else:
        disp_name = input(f"  Bezeichnung [{shift_key}]: ").strip() or shift_key
        start     = input( "  Startzeit (HH:MM)     [07:15]: ").strip() or "07:15"
        end       = input( "  Endzeit   (HH:MM)     [15:45]: ").strip() or "15:45"
        next_d    = input( "  Endet am Folgetag?    (j/n) [n]: ").strip().lower() == 'j'

        new_def = {'name': disp_name, 'start': start, 'end': end}
        if next_d:
            new_def['next_day'] = True

        lines = [
            f'\n  "{shift_key}":',
            f'    name: "{disp_name}"',
            f'    start: "{start}"',
            f'    end: "{end}"',
        ]
        if next_d:
            lines.append('    next_day: true')
        yaml_block = '\n'.join(lines) + '\n'
        label = f"{start}-{end}" + (" (+1 Tag)" if next_d else "")

    # In Laufzeit-Config eintragen
    config.setdefault('shifts', {})[shift_key] = new_def

    # An config.yaml anhaengen
    with open(config_path, 'a', encoding='utf-8') as f:
        f.write(yaml_block)

    print(f"  Gespeichert: '{shift_key}' -> {label}")
    return True


def serve_ics(ics_path: Path, port: int = 8765):
    """
    Startet einen temporären HTTP-Server damit das iPhone die .ics Datei
    direkt in Safari öffnen kann. iPhone und PC müssen im gleichen WLAN sein.
    """
    # Lokale IP ermitteln
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
    finally:
        s.close()

    http_url  = f"http://{local_ip}:{port}/{ics_path.name}"
    webcal_url = f"webcal://{local_ip}:{port}/{ics_path.name}"

    class Handler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=str(ics_path.parent), **kwargs)
        def log_message(self, *args):
            pass

    server = http.server.HTTPServer(('', port), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()

    print()
    print("  ============================================")
    print("  iPhone-Import via Safari")
    print("  ============================================")
    print()
    print("  Voraussetzung: iPhone im gleichen WLAN wie dieser PC")
    print()
    print("  Schritt 1 - Safari oeffnen und DIESE Adresse eingeben:")
    print()
    print(f"     {webcal_url}")
    print()
    print("  --> Kalender-App oeffnet sich automatisch")
    print()
    print("  Falls das nicht klappt, Alternative mit http://:")
    print(f"     {http_url}")
    print("  --> Datei wird heruntergeladen, dann antippen")
    print()
    print("  Fuer Kollegen: gleiche URL, sie brauchen nur Safari")
    print()
    print("  Enter druecken zum Beenden des Servers...")
    input()
    server.shutdown()


def push_config(config_path: Path, new_keys: list):
    """Committet und pusht config.yaml mit den neu hinzugefügten Diensten."""
    repo_dir = config_path.parent
    keys_str = ', '.join(new_keys)
    try:
        subprocess.run(['git', 'add', str(config_path)], cwd=repo_dir, check=True)
        subprocess.run(
            ['git', 'commit', '-m', f'config: neue Dienste ergaenzt ({keys_str})'],
            cwd=repo_dir, check=True
        )
        subprocess.run(['git', 'push'], cwd=repo_dir, check=True)
        print(f"\n  config.yaml gepusht ({keys_str})")
    except subprocess.CalledProcessError as e:
        print(f"\n  Hinweis: Git-Push fehlgeschlagen ({e}) – config.yaml wurde lokal gespeichert.")


def name_to_filename(name: str) -> str:
    """'Mangels, Nils' -> 'Mangels_Nils'"""
    return re.sub(r'[^\w]', '_', name).strip('_')


def save_ics(file_path: Path, name: str, month: int, year: int,
             schedule: dict, config: dict, output_dir: Path,
             prefix: str = 'Dienstplan') -> Path:
    MONATE = {1:'Januar',2:'Februar',3:'März',4:'April',5:'Mai',6:'Juni',
              7:'Juli',8:'August',9:'September',10:'Oktober',11:'November',12:'Dezember'}

    cal = Calendar()
    cal.add('prodid', '-//Dienstplan Kalender//DE')
    cal.add('version', '2.0')
    cal.add('calscale', 'GREGORIAN')
    cal.add('method', 'PUBLISH')
    cal.add('x-wr-calname', f'{prefix} {name}')

    event_count = 0
    for day in sorted(schedule.keys()):
        for event in create_events(day, month, year, schedule[day], config):
            cal.add_component(event)
            event_count += 1

    ical_data = cal.to_ical()
    base = f"{prefix}_{name_to_filename(name)}_{MONATE[month]}{year}"

    for ext in ('.ical', '.ics'):
        with open(output_dir / (base + ext), 'wb') as f:
            f.write(ical_data)

    return output_dir / (base + '.ics'), event_count


def save_single_ics_files(file_path: Path, name: str, month: int, year: int,
                          schedule: dict, config: dict, output_dir: Path,
                          prefix: str = 'Dienstplan') -> tuple:
    MONATE = {1:'Januar',2:'Februar',3:'März',4:'April',5:'Mai',6:'Juni',
              7:'Juli',8:'August',9:'September',10:'Oktober',11:'November',12:'Dezember'}

    folder_name = f"{prefix}_{name_to_filename(name)}_{MONATE[month]}{year}_einzeln"
    folder = output_dir / folder_name
    folder.mkdir(exist_ok=True)

    count = 0
    for day in sorted(schedule.keys()):
        for evt in create_events(day, month, year, schedule[day], config):
            cal = Calendar()
            cal.add('prodid', '-//Dienstplan Kalender//DE')
            cal.add('version', '2.0')
            cal.add('calscale', 'GREGORIAN')
            cal.add('method', 'PUBLISH')
            cal.add_component(evt)

            title = str(evt.get('summary'))
            safe_title = re.sub(r'[^\w]', '_', title).strip('_')
            fname = f"{day:02d}_{safe_title}"
            for ext in ('.ical', '.ics'):
                with open(folder / (fname + ext), 'wb') as f:
                    f.write(cal.to_ical())
            count += 1

    return folder, count


def save_json(file_path: Path, name: str, month: int, year: int,
              schedule: dict, config: dict, output_dir: Path,
              prefix: str = 'Dienstplan') -> Path:
    MONATE = {1:'Januar',2:'Februar',3:'März',4:'April',5:'Mai',6:'Juni',
              7:'Juli',8:'August',9:'September',10:'Oktober',11:'November',12:'Dezember'}

    events = []
    for day in sorted(schedule.keys()):
        for evt in create_events(day, month, year, schedule[day], config):
            dt_start = evt.get('dtstart').dt
            dt_end   = evt.get('dtend').dt
            events.append({
                'titel': str(evt.get('summary')),
                'start': dt_start.strftime('%Y-%m-%dT%H:%M:%S'),
                'ende':  dt_end.strftime('%Y-%m-%dT%H:%M:%S'),
            })

    data = {
        'name':   name,
        'monat':  f"{MONATE[month]} {year}",
        'events': events,
    }

    base = f"{prefix}_{name_to_filename(name)}_{MONATE[month]}{year}"
    json_path = output_dir / (base + '.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_path


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    debug = '--debug' in sys.argv
    alle  = '--alle'  in sys.argv
    scan  = '--scan'  in sys.argv
    serve = '--serve' in sys.argv

    # --serve: direkt eine .ics Datei bereitstellen (kein PDF/Excel nötig)
    if serve and args:
        serve_ics(Path(args[0]))
        sys.exit(0)

    if not args:
        print("Verwendung: python dienstplan.py <datei.pdf|xlsx> [\"Name\"] [--alle] [--debug]")
        sys.exit(1)

    file_path = Path(args[0])
    config_path = Path(__file__).parent / 'config.yaml'

    if not file_path.exists():
        print(f"Fehler: Datei nicht gefunden: {file_path}")
        sys.exit(1)

    if not config_path.exists():
        print(f"Fehler: config.yaml nicht gefunden in {config_path.parent}")
        sys.exit(1)

    config = load_config(config_path)

    is_excel = file_path.suffix.lower() in {'.xlsx', '.xls', '.xlsm'}

    if scan:
        if not is_excel:
            print("--scan funktioniert nur mit Excel-Dateien.")
            sys.exit(1)
        scan_excel(file_path)
        sys.exit(0)
    extract_fn     = extract_schedule_excel if is_excel else extract_schedule
    find_names_fn  = find_all_names_excel   if is_excel else find_all_names
    filetype_label = "Excel" if is_excel else "PDF"
    filename_prefix = "Kreuzchenplan" if is_excel else "Gesamtdienstplan"

    # --- Modus: alle KollegInnen ---
    if alle:
        print(f"{filetype_label}: {file_path.name}")
        print("Suche alle Namen...\n")

        names = find_names_fn(file_path)
        if not names:
            print("Keine Namen gefunden.")
            sys.exit(1)

        print(f"{len(names)} Personen gefunden:")
        for n in names:
            print(f"  - {n}")
        print()

        output_dir = file_path.parent

        # Alle unbekannten Kuerzel vorab sammeln und einmalig abfragen
        print("Pruefe auf unbekannte Kuerzel...")
        all_unknown = set()
        all_schedules = {}
        for n in names:
            sched, month, year = extract_fn(file_path, n)
            all_schedules[n] = (sched, month, year)
            if sched:
                for key in find_unknown_shifts(sched, config):
                    all_unknown.add(key)

        new_keys = []
        if all_unknown:
            print(f"\n{len(all_unknown)} unbekannte(s) Kuerzel gefunden.\n")
            for key in sorted(all_unknown):
                ask_and_save_shift(key, config_path, config)
                new_keys.append(key)
        if new_keys:
            push_config(config_path, new_keys)

        print()
        ok = skip = 0
        for n in names:
            sched, month, year = all_schedules[n]
            if not sched or not month:
                print(f"  {n}: keine Dienste, uebersprungen")
                skip += 1
                continue
            output_path, count = save_ics(file_path, n, month, year, sched, config, output_dir, filename_prefix)
            save_json(file_path, n, month, year, sched, config, output_dir, filename_prefix)
            save_single_ics_files(file_path, n, month, year, sched, config, output_dir, filename_prefix)
            print(f"  {n}: {count} Eintraege -> {output_path.name}")
            ok += 1

        print(f"\n{ok} Dateien erstellt, {skip} uebersprungen.")
        print(f"Ordner: {output_dir}")
        sys.exit(0)

    # --- Modus: einzelne Person ---
    name = args[1] if len(args) > 1 else config.get('name', '')
    if not name:
        print("Fehler: Kein Name angegeben.")
        sys.exit(1)

    print(f"Suche nach: {name}")
    print(f"{filetype_label}: {file_path.name}")

    schedule, month, year = extract_fn(file_path, name, debug)

    if not schedule:
        print("\nKeine Dienste gefunden!")
        if is_excel:
            print("Tipps:")
            print("  - Name exakt wie in der Excel-Datei (z.B. 'Nachname, Vorname')")
            print("  - Mit --debug starten fuer Details (zeigt gefundene Namen)")
        else:
            print("Tipps:")
            print("  - Name exakt wie im PDF (z.B. 'Nachname, Vorname')")
            print("  - Mit --debug starten fuer Details")
        sys.exit(1)

    if not month or not year:
        print("Fehler: Monat/Jahr konnte nicht gelesen werden.")
        sys.exit(1)

    # Unbekannte Kuerzel abfragen
    unknown  = find_unknown_shifts(schedule, config)
    new_keys = []
    if unknown:
        print(f"\n{len(unknown)} unbekannte(s) Kuerzel gefunden:")
        for key in unknown:
            ask_and_save_shift(key, config_path, config)
            new_keys.append(key)
    if new_keys:
        push_config(config_path, new_keys)

    MONATE = {1:'Januar', 2:'Februar', 3:'März',   4:'April',
              5:'Mai',    6:'Juni',    7:'Juli',    8:'August',
              9:'September', 10:'Oktober', 11:'November', 12:'Dezember'}
    print(f"\n{MONATE[month]} {year} — {len(schedule)} Tag(e) mit Dienst:\n")

    for day in sorted(schedule.keys()):
        info   = schedule[day]
        events = create_events(day, month, year, info, config)
        wd     = info.get('weekday', '   ')
        marker = ' (uebersprungen)' if not events else ''
        print(f"  {day:2d}. {wd}  {info['shift']}{marker}")

    output_path, event_count = save_ics(
        file_path, name, month, year, schedule, config, file_path.parent, filename_prefix
    )
    json_path = save_json(
        file_path, name, month, year, schedule, config, file_path.parent, filename_prefix
    )
    einzeln_folder, einzeln_count = save_single_ics_files(
        file_path, name, month, year, schedule, config, file_path.parent, filename_prefix
    )
    print(f"\n{event_count} Eintraege erstellt")
    print(f"Datei (gesamt):   {output_path}")
    print(f"Einzeldateien:    {einzeln_folder} ({einzeln_count} Dateien)")
    print(f"JSON (Shortcuts): {json_path}")


if __name__ == '__main__':
    main()
